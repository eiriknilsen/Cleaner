import gooeypie as gp
import os
import pandas as pd
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def backup_file(file_path):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = file_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
    shutil.copyfile(file_path, backup_path)
    print(f"Backup saved as {backup_path}")


def create_new_sheet(workbook, sheet_name):
    """Create a new sheet in the Excel file and populate it with data."""
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        print(f"Sheet '{sheet_name}' created.")

        # Populate the new sheet with data
        data = [
            ['Arknavn', 'Duplikat varenr', 'Mangler basisenhet', 'Antall pakker', 'Levnr mangler lev.'],
            ['Logistikk Advanced'],
            ['VareAttributter'],
            ['Parti og serienummer profiler'],
            ['Måleenhete_Konverteringsenheter'],
            ['Vareprofiler'],
            ['Vareprisgrupper'],
            ['Lokasjoner'],
            ['Kryssreferanser']
        ]

        for i, row in enumerate(data, start=1):
            for j, value in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=value)

        # Create a table from the data
        table = Table(displayName="AnalyseTable", ref=f"A1:E{len(data)}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)
    else:
        print(f"Sheet '{sheet_name}' already exists.")


def clean_sheet_1(df):
    """Clean the data in the Logistikk Advanced sheet and perform specified checks."""
    df.drop_duplicates(inplace=True)
    df.columns = [col.lower() for col in df.columns if isinstance(col, str)]

    # Count duplicates in 'Varenummer'
    duplicate_count = df.duplicated(subset=['varenummer'], keep=False).sum()

    # Count TRUE values in 'Pakke?'
    true_count = df['pakke?'].sum() if 'pakke?' in df.columns else 0

    # Count missing values in 'Basis_Enhet'
    missing_values_count = df['basis_enhet'].isna().sum() if 'basis_enhet' in df.columns else 0

    # Check if 'Lev Varenr' has a value and corresponding 'Leverandørnr' value
    lev_varnr_has_value = df['lev varenr'].notna() if 'lev varenr' in df.columns else None
    leverandornr_value_missing = lev_varnr_has_value & df[
        'leverandornr'].isna() if 'leverandornr' in df.columns else None
    lev_value_mismatch_count = leverandornr_value_missing.sum() if leverandornr_value_missing is not None else 0

    return df, duplicate_count, true_count, missing_values_count, lev_value_mismatch_count


def clean_excel(file_path):
    backup_file(file_path)  # Save a backup before making changes

    # Load the Excel file
    xls = pd.ExcelFile(file_path)
    print(xls.sheet_names)  # Debugging

    workbook = load_workbook(file_path)

    # Create a new sheet named 'Analyse'
    create_new_sheet(workbook, 'Analyse')

    duplicate_count = 0
    true_count = 0
    missing_values_count = 0
    lev_value_mismatch_count = 0

    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            print(df.head())  # Debugging

            if sheet_name == "Logistikk Advanced":
                # Perform cleaning operations
                df, duplicate_count, true_count, missing_values_count, lev_value_mismatch_count = clean_sheet_1(df)

            # Save the cleaned sheet back to Excel
            with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        except Exception as e:
            print(f"Error reading sheet {sheet_name}: {e}")

    # Write the counts to Analyse sheet
    sheet = workbook["Analyse"]
    sheet["B2"] = duplicate_count
    sheet["C2"] = missing_values_count
    sheet["D2"] = true_count
    sheet["E2"] = lev_value_mismatch_count

    # Save the workbook with the new sheet
    workbook.save(file_path)


def clean_file(event):
    file_path = file_input.text.strip('"')  # Remove double quotes if present
    file_path = file_path.strip()  # Remove any leading/trailing whitespace
    if file_path and os.path.isfile(file_path) and file_path.endswith('.xlsx'):
        file_name = os.path.basename(file_path)
        clean_excel(file_path)
        result_lbl.text = "File cleaned: " + file_name
    else:
        result_lbl.text = "Invalid file path. Please enter a valid Excel file path."


# Create the GUI
app = gp.GooeyPieApp('Logistikk Advanced Cleaner')
app.width = 1000

# Create the label, input field, and button
label = gp.Label(app, 'Input')
file_input = gp.Input(app)
file_btn = gp.Button(app, 'Clean File', clean_file)
result_lbl = gp.Label(app, '')
file_input.width = 100

# Add the components to the app
app.set_grid(4, 1)
app.add(label, 1, 1, align='center')
app.add(file_input, 2, 1, align='center')
app.add(file_btn, 3, 1, align='center')
app.add(result_lbl, 4, 1, align='center')

# Run the app
app.run()
